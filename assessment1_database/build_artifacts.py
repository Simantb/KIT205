from __future__ import annotations

import csv
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from reportlab.lib import colors
from reportlab.lib.pagesizes import A1, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import Paragraph
from reportlab.pdfgen import canvas


ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = ROOT / "output"
RESULTS_CSV = ROOT / "evaluation_results.csv"
DEVLOG_CSV = ROOT / "devlog_sample.csv"
POSTER_PDF = OUTPUT_DIR / "assessment1_database_poster_a1.pdf"
DEVLOG_XLSX = OUTPUT_DIR / "assessment1_database_devlog_draft.xlsx"
LOGO_PATH = Path("/Users/simantbhattarai/Desktop/sem 2026/Data structure /Content/images/UTAS Logo.png")
REPO_LINK = "https://github.com/Simantb/KIT205/tree/main/assessment1_database"
AUTHOR_NAME = "Simant Bhattarai"


def read_results() -> list[dict[str, float]]:
    rows: list[dict[str, float]] = []
    with RESULTS_CSV.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            rows.append(
                {
                    "relationships": float(row["relationships"]),
                    "p1_insert": float(row["prototype1_insert_ms"]),
                    "p2_insert": float(row["prototype2_insert_ms"]),
                    "p1_lookup": float(row["prototype1_lookup_ms"]),
                    "p2_lookup": float(row["prototype2_lookup_ms"]),
                }
            )
    if not rows:
        raise ValueError("No benchmark rows found in evaluation_results.csv")
    return rows


def read_devlog_rows() -> list[list[str]]:
    with DEVLOG_CSV.open(newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        return [row for row in reader]


def build_devlog_workbook(rows: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "DevLog"

    heading_fill = PatternFill("solid", fgColor="7B0C18")
    subheading_fill = PatternFill("solid", fgColor="E9D8D6")
    body_fill = PatternFill("solid", fgColor="FBF8F6")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="B8A9A5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:F1")
    ws["A1"] = "Assessment Task 1 DevLog Draft"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].fill = heading_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        "This workbook is a structured draft based on the completed implementation. "
        "If your unit provides a compulsory template, copy these entries into that template."
    )
    ws["A2"].fill = body_fill
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 36

    header = rows[0]
    for col, value in enumerate(header, start=1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.font = white_font
        cell.fill = heading_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_index, row in enumerate(rows[1:], start=5):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            cell.fill = body_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        1: 14,
        2: 32,
        3: 48,
        4: 44,
        5: 28,
        6: 18,
    }
    for column_index, width in widths.items():
        column_letter = get_column_letter(column_index)
        ws.column_dimensions[column_letter] = ColumnDimension(ws, min=column_index, max=column_index, width=width)

    ws.freeze_panes = "A5"

    summary = wb.create_sheet("Milestones")
    summary["A1"] = "Milestone"
    summary["B1"] = "Purpose"
    summary["C1"] = "Evidence in Project"
    for cell in summary[1]:
        cell.font = white_font
        cell.fill = heading_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    milestone_rows = [
        (
            "Initial commit",
            "Establish the domain, relationship, and project structure.",
            "README.md topic selection and starter project files.",
        ),
        (
            "Prototype 1 complete",
            "Implement the linked-list + AVL prototype and prove correctness.",
            "prototype1.c/.h plus unit tests in tests.c.",
        ),
        (
            "Prototype 2 complete",
            "Implement the hash-table + array prototype and prove correctness.",
            "prototype2.c/.h plus shared tests in tests.c.",
        ),
        (
            "Database implementation",
            "Provide complete operations and example outputs for both designs.",
            "main.c example output and shared record-management logic.",
        ),
        (
            "Evaluation code complete",
            "Benchmark the time-critical operations and capture results for the poster.",
            "evaluation.c, evaluation_results.csv, and poster artifacts.",
        ),
    ]
    for row_index, values in enumerate(milestone_rows, start=2):
        for col_index, value in enumerate(values, start=1):
            cell = summary.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = body_fill if row_index % 2 == 0 else subheading_fill

    summary.column_dimensions["A"].width = 25
    summary.column_dimensions["B"].width = 42
    summary.column_dimensions["C"].width = 54

    checklist = wb.create_sheet("Submission Checklist")
    checklist["A1"] = "Item"
    checklist["B1"] = "Status"
    checklist["C1"] = "Notes"
    for cell in checklist[1]:
        cell.font = white_font
        cell.fill = heading_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    checklist_rows = [
        ("Two alternative prototypes in C", "Complete", "Both prototypes compile and run in the current package."),
        ("Unit tests", "Complete", "14 tests currently pass."),
        ("Performance evaluation", "Complete", "Results written to evaluation_results.csv."),
        ("Research poster", "Complete", "A1 PDF poster generated in output/."),
        ("DevLog workbook", "Draft complete", "Use this content directly or transfer into the official template if required."),
        ("GitHub milestone history", "Needs honest finalisation", "Current commits can be added now, but earlier real-time history cannot be fabricated."),
    ]
    for row_index, values in enumerate(checklist_rows, start=2):
        for col_index, value in enumerate(values, start=1):
            cell = checklist.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = body_fill if row_index % 2 == 0 else subheading_fill

    checklist.column_dimensions["A"].width = 30
    checklist.column_dimensions["B"].width = 18
    checklist.column_dimensions["C"].width = 58

    wb.save(DEVLOG_XLSX)


def make_paragraph_style(name: str, size: int, leading: int, color: colors.Color, bold: bool = False) -> ParagraphStyle:
    styles = getSampleStyleSheet()
    base = styles["BodyText"]
    font_name = "Helvetica-Bold" if bold else "Helvetica"
    return ParagraphStyle(
        name=name,
        parent=base,
        fontName=font_name,
        fontSize=size,
        leading=leading,
        textColor=color,
        spaceAfter=0,
        spaceBefore=0,
    )


def draw_paragraph(
    c: canvas.Canvas,
    text: str,
    x: float,
    y: float,
    width: float,
    height: float,
    style: ParagraphStyle,
) -> None:
    paragraph = Paragraph(text, style)
    needed_width, needed_height = paragraph.wrap(width, height)
    paragraph.drawOn(c, x, y + height - needed_height)


def draw_section_box(
    c: canvas.Canvas,
    title: str,
    body: str,
    x: float,
    y: float,
    width: float,
    height: float,
    title_fill: colors.Color,
    body_fill: colors.Color,
) -> None:
    c.setFillColor(body_fill)
    c.roundRect(x, y, width, height, 16, fill=1, stroke=0)
    c.setFillColor(title_fill)
    c.roundRect(x, y + height - 38, width, 38, 16, fill=1, stroke=0)
    c.rect(x, y + height - 19, width, 19, fill=1, stroke=0)

    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(x + 14, y + height - 24, title)

    body_style = make_paragraph_style("body", 12, 16, colors.HexColor("#22303C"))
    draw_paragraph(c, body, x + 14, y + 14, width - 28, height - 56, body_style)


def draw_small_table(
    c: canvas.Canvas,
    x: float,
    y: float,
    width: float,
    row_height: float,
    headers: list[str],
    rows: list[list[str]],
) -> None:
    header_fill = colors.HexColor("#7B0C18")
    alt_fill = colors.HexColor("#FBF7F5")
    grid = colors.HexColor("#C7B8B4")
    text_color = colors.HexColor("#22303C")
    col_count = len(headers)
    col_width = width / col_count
    total_height = row_height * (len(rows) + 1)

    c.setStrokeColor(grid)
    c.setLineWidth(0.8)

    c.setFillColor(header_fill)
    c.rect(x, y + total_height - row_height, width, row_height, fill=1, stroke=0)

    for row_index, row in enumerate(rows):
        if row_index % 2 == 0:
            c.setFillColor(alt_fill)
            c.rect(x, y + total_height - row_height * (row_index + 2), width, row_height, fill=1, stroke=0)

    for i in range(col_count + 1):
        line_x = x + i * col_width
        c.line(line_x, y, line_x, y + total_height)
    for i in range(len(rows) + 2):
        line_y = y + i * row_height
        c.line(x, line_y, x + width, line_y)

    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(colors.white)
    for index, header in enumerate(headers):
        text_x = x + index * col_width + 6
        text_y = y + total_height - row_height + 8
        c.drawString(text_x, text_y, header)

    c.setFont("Helvetica", 10)
    c.setFillColor(text_color)
    for row_index, row in enumerate(rows):
        for col_index, value in enumerate(row):
            text_x = x + col_index * col_width + 6
            text_y = y + total_height - row_height * (row_index + 2) + 8
            c.drawString(text_x, text_y, value)


def draw_diagram_box(c: canvas.Canvas, x: float, y: float, width: float, height: float, title: str, items: list[str], accent: colors.Color) -> None:
    c.setFillColor(colors.white)
    c.roundRect(x, y, width, height, 14, fill=1, stroke=0)
    c.setStrokeColor(accent)
    c.setLineWidth(2)
    c.roundRect(x, y, width, height, 14, fill=0, stroke=1)
    c.setFillColor(accent)
    c.setFont("Helvetica-Bold", 15)
    c.drawString(x + 12, y + height - 24, title)
    c.setFillColor(colors.HexColor("#22303C"))
    c.setFont("Helvetica", 12)
    bullet_y = y + height - 48
    for item in items:
        c.circle(x + 18, bullet_y + 4, 2.2, fill=1, stroke=0)
        c.drawString(x + 28, bullet_y, item)
        bullet_y -= 18


def draw_log_chart(
    c: canvas.Canvas,
    title: str,
    x: float,
    y: float,
    width: float,
    height: float,
    x_values: list[float],
    series: list[tuple[str, list[float], colors.Color]],
) -> None:
    bg = colors.white
    grid = colors.HexColor("#D7CCCA")
    axis = colors.HexColor("#5B6770")
    c.setFillColor(bg)
    c.roundRect(x, y, width, height, 14, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor("#BDAFAC"))
    c.setLineWidth(1)
    c.roundRect(x, y, width, height, 14, fill=0, stroke=1)

    c.setFillColor(colors.HexColor("#22303C"))
    c.setFont("Helvetica-Bold", 15)
    c.drawString(x + 14, y + height - 24, title)
    c.setFont("Helvetica", 10)
    c.setFillColor(colors.HexColor("#52606D"))
    c.drawString(x + 14, y + height - 38, "Log scale (milliseconds)")

    chart_left = x + 56
    chart_bottom = y + 42
    chart_width = width - 80
    chart_height = height - 84

    all_values = [value for _, values, _ in series for value in values]
    min_power = math.floor(math.log10(min(all_values)))
    max_power = math.ceil(math.log10(max(all_values)))

    c.setStrokeColor(grid)
    c.setLineWidth(0.8)
    for power in range(min_power, max_power + 1):
        value = 10 ** power
        ratio = (math.log10(value) - min_power) / (max_power - min_power)
        y_pos = chart_bottom + chart_height * ratio
        c.line(chart_left, y_pos, chart_left + chart_width, y_pos)
        c.setFillColor(axis)
        c.setFont("Helvetica", 9)
        c.drawRightString(chart_left - 6, y_pos - 3, f"{value:g}")

    min_x = min(x_values)
    max_x = max(x_values)
    c.setStrokeColor(axis)
    c.setLineWidth(1.1)
    c.line(chart_left, chart_bottom, chart_left, chart_bottom + chart_height)
    c.line(chart_left, chart_bottom, chart_left + chart_width, chart_bottom)

    for x_value in x_values:
        ratio = (x_value - min_x) / (max_x - min_x)
        x_pos = chart_left + chart_width * ratio
        c.setStrokeColor(grid)
        c.line(x_pos, chart_bottom, x_pos, chart_bottom + chart_height)
        c.setFillColor(axis)
        c.setFont("Helvetica", 9)
        c.drawCentredString(x_pos, chart_bottom - 14, f"{int(x_value):,}")

    def point_for(x_value: float, y_value: float) -> tuple[float, float]:
        x_ratio = (x_value - min_x) / (max_x - min_x)
        y_ratio = (math.log10(y_value) - min_power) / (max_power - min_power)
        return chart_left + chart_width * x_ratio, chart_bottom + chart_height * y_ratio

    for label, values, colour in series:
        c.setStrokeColor(colour)
        c.setFillColor(colour)
        c.setLineWidth(2.5)
        points = [point_for(xv, yv) for xv, yv in zip(x_values, values)]
        for i in range(len(points) - 1):
            c.line(points[i][0], points[i][1], points[i + 1][0], points[i + 1][1])
        for px, py in points:
            c.circle(px, py, 3, fill=1, stroke=0)

    legend_x = x + 18
    legend_y = y + 12
    for label, _, colour in series:
        c.setFillColor(colour)
        c.rect(legend_x, legend_y, 12, 12, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#22303C"))
        c.setFont("Helvetica", 10)
        c.drawString(legend_x + 18, legend_y + 1, label)
        legend_x += stringWidth(label, "Helvetica", 10) + 48


def build_poster(results: list[dict[str, float]]) -> None:
    page_width, page_height = landscape(A1)
    c = canvas.Canvas(str(POSTER_PDF), pagesize=(page_width, page_height))

    background = colors.HexColor("#F4EFEC")
    card_fill = colors.HexColor("#FFFDFC")
    primary = colors.HexColor("#7B0C18")
    secondary = colors.HexColor("#D98F7C")
    ink = colors.HexColor("#22303C")

    c.setFillColor(background)
    c.rect(0, 0, page_width, page_height, fill=1, stroke=0)

    c.setFillColor(primary)
    c.rect(0, page_height - 96, page_width, 96, fill=1, stroke=0)
    c.setFillColor(colors.HexColor("#EAD6D3"))
    c.rect(0, page_height - 112, page_width, 16, fill=1, stroke=0)

    if LOGO_PATH.exists():
        logo = ImageReader(str(LOGO_PATH))
        c.drawImage(logo, page_width - 285, page_height - 82, width=235, height=46, preserveAspectRatio=True, mask="auto")

    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 32)
    c.drawString(46, page_height - 48, "Comparing Two C Database Prototypes for Large Publishing Networks")
    c.setFont("Helvetica", 17)
    c.drawString(46, page_height - 74, f"{AUTHOR_NAME}  |  KIT205 Assessment Task 1  |  {REPO_LINK}")

    summary_x = 46
    summary_y = page_height - 176
    summary_w = page_width - 92
    summary_h = 54
    c.setFillColor(colors.white)
    c.roundRect(summary_x, summary_y, summary_w, summary_h, 14, fill=1, stroke=0)
    c.setFillColor(ink)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(summary_x + 18, summary_y + 33, "Verified package status")
    c.setFont("Helvetica", 13)
    c.drawString(
        summary_x + 18,
        summary_y + 14,
        "Both prototypes compiled and ran successfully on 18 May 2026. Unit tests: 14 passed, 0 failed.",
    )

    margin = 46
    gutter = 20
    col1 = 520
    col2 = 640
    col3 = page_width - margin * 2 - gutter * 2 - col1 - col2
    usable_height = page_height - 270
    top_y = page_height - 250

    left_x = margin
    middle_x = left_x + col1 + gutter
    right_x = middle_x + col2 + gutter

    draw_section_box(
        c,
        "Introduction",
        (
            "This project prototypes a database-style system for a publishing catalogue. "
            "The data model is many-to-many: one <b>book</b> can have many <b>authors</b>, "
            "and one author can contribute to many books. Large digital catalogues need this "
            "relationship to stay correct, ordered, and fast at scale."
        ),
        left_x,
        top_y - 200,
        col1,
        200,
        primary,
        card_fill,
    )

    draw_section_box(
        c,
        "Required Operations",
        (
            "1. Insert a book-author relationship.<br/>"
            "2. Print an ordered list of all books.<br/>"
            "3. Print an ordered list of all authors for a given book.<br/>"
            "4. Print an ordered list of all books for a given author."
        ),
        left_x,
        top_y - 390,
        col1,
        170,
        primary,
        card_fill,
    )

    draw_section_box(
        c,
        "Methodology",
        (
            "The program generates simulated book-author relationships at five scales: "
            "1,000, 3,000, 6,000, 12,000, and 24,000 relationships. Each prototype processes "
            "the same generated data. Two time-critical operations are measured: total insertion time "
            "and repeated author-lookups by book ID."
        ),
        left_x,
        top_y - 590,
        col1,
        180,
        primary,
        card_fill,
    )

    draw_section_box(
        c,
        "Conclusion",
        (
            "Prototype 2 is the stronger recommendation for very large datasets in this domain. "
            "Its hash-table record lookup keeps both insertion and query time substantially lower "
            "than Prototype 1 as the database grows. Prototype 1 still provides a correct ordered "
            "tree-based design, but linked-list record search becomes the dominant bottleneck."
        ),
        left_x,
        top_y - 790,
        col1,
        180,
        primary,
        card_fill,
    )

    draw_diagram_box(
        c,
        middle_x,
        top_y - 240,
        col2 / 2 - 10,
        240,
        "Prototype 1",
        [
            "Sorted linked list of books",
            "Sorted linked list of authors",
            "AVL tree of authors per book",
            "AVL tree of books per author",
        ],
        primary,
    )
    draw_diagram_box(
        c,
        middle_x + col2 / 2 + 10,
        top_y - 240,
        col2 / 2 - 10,
        240,
        "Prototype 2",
        [
            "Hash table of books",
            "Hash table of authors",
            "Sorted dynamic array per book",
            "Sorted dynamic array per author",
        ],
        secondary,
    )

    draw_section_box(
        c,
        "Why These Structures?",
        (
            "Prototype 1 combines linked lists and balanced trees to preserve order naturally, "
            "but record discovery still requires linear traversal. Prototype 2 trades tree-based ordering "
            "for average-case constant-time hash lookup, while using sorted arrays to preserve ordered output."
        ),
        middle_x,
        top_y - 430,
        col2,
        160,
        primary,
        card_fill,
    )

    complexity_headers = ["Operation", "Prototype 1", "Prototype 2"]
    complexity_rows = [
        ["Insert relationship", "O(B + A + log r)", "O(1 avg. lookup + r shift)"],
        ["List all books", "O(B)", "O(B log B)"],
        ["Authors for a book", "O(B + r)", "O(1 avg. + r)"],
        ["Books for an author", "O(A + r)", "O(1 avg. + r)"],
    ]
    draw_small_table(c, middle_x, top_y - 650, col2, 28, complexity_headers, complexity_rows)

    draw_section_box(
        c,
        "References",
        (
            "KIT205 lecture and tutorial materials; C standard library documentation; "
            "algorithm notes for AVL trees and hash tables; GitHub repository for version control; "
            "AI support used for drafting and implementation assistance."
        ),
        middle_x,
        top_y - 820,
        col2,
        138,
        primary,
        card_fill,
    )

    x_values = [row["relationships"] for row in results]
    insert_series = [
        ("Prototype 1 insert", [row["p1_insert"] for row in results], primary),
        ("Prototype 2 insert", [row["p2_insert"] for row in results], secondary),
    ]
    lookup_series = [
        ("Prototype 1 lookup", [row["p1_lookup"] for row in results], colors.HexColor("#275D8A")),
        ("Prototype 2 lookup", [row["p2_lookup"] for row in results], colors.HexColor("#2F8F6B")),
    ]

    draw_log_chart(c, "Insertion Performance", right_x, top_y - 330, col3, 330, x_values, insert_series)
    draw_log_chart(c, "Lookup Performance", right_x, top_y - 690, col3, 330, x_values, lookup_series)

    result_headers = ["Relationships", "P1 Insert", "P2 Insert", "P1 Lookup", "P2 Lookup"]
    result_rows = [
        [
            f"{int(row['relationships']):,}",
            f"{row['p1_insert']:.3f}",
            f"{row['p2_insert']:.3f}",
            f"{row['p1_lookup']:.3f}",
            f"{row['p2_lookup']:.3f}",
        ]
        for row in results
    ]
    draw_small_table(c, right_x, top_y - 930, col3, 28, result_headers, result_rows)

    c.setFont("Helvetica", 10)
    c.setFillColor(colors.HexColor("#52606D"))
    c.drawRightString(page_width - 46, 24, "Generated from build_artifacts.py using verified evaluation_results.csv")
    c.save()


def main() -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    results = read_results()
    devlog_rows = read_devlog_rows()
    build_devlog_workbook(devlog_rows)
    build_poster(results)
    print(f"Wrote {POSTER_PDF}")
    print(f"Wrote {DEVLOG_XLSX}")


if __name__ == "__main__":
    main()
